from openpyxl import workbook
import openpyxl

# function to get number of rows in excel 
def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet= workbook.get_sheet_by_name(sheetName)
    return(sheet.max_row)

# function to get number of columns in excel
def getColCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet= workbook.get_sheet_by_name(sheetName)
    return(sheet.max_column)

# function to read data from excel
def readData(file,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet= workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum,column=columnno).value

# function to write data in excel
def writeData(file,sheetName,rownum,columnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet= workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rownum,column=columnno).value=data
    workbook.save(file)